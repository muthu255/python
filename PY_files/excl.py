from openpyxl import *
from openpyxl import Workbook
from openpyxl import load_workbook



wb = load_workbook('muthu.xlsx')
sheetlist = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheetlist[0])


"""
ws = wb.active
ws1 = wb.create_sheet("Metadata") # insert at the end (default)
ws2 = wb.create_sheet("Count", 0) # insert at first position
ws3 = wb.create_sheet("S-T", 1) # insert at the penultimate position
ws4 = wb.create_sheet("T-S", 2)
ws5 = wb.create_sheet("Not null", 3)
ws.title='Results'
wb.save('muthu.xlsx')
"""









